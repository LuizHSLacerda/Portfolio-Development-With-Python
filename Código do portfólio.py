import tkinter as tk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Função para buscar a previsão do tempo
def buscar_previsao():
    # Inicializa o driver
    driver = webdriver.Chrome()

    # Abre o Google e busca pela previsão do tempo
    driver.get("https://www.google.com")
    caixa_busca = driver.find_element(By.NAME, "q")
    caixa_busca.send_keys("previsão do tempo")
    caixa_busca.send_keys(Keys.RETURN)

    # Espera a página carregar e coleta os dados
    time.sleep(3)  # Adapte conforme a velocidade da sua conexão

    # Captura a temperatura e umidade do ar
    temperatura = driver.find_element(By.XPATH, '//*[@id="wob_tm"]').text
    umidade = driver.find_element(By.XPATH, '//*[@id="wob_hm"]').text

    # Fecha o driver
    driver.quit()

    # Define o caminho do arquivo
    caminho_arquivo = r"C:\Users\lusya\OneDrive\Área de Trabalho\Portfólio Development With Python\historico_previsao_tempo.xlsx"

    # Carrega ou cria uma planilha Excel
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data e Hora", "Temperatura", "Umidade"])

    # Adiciona os dados coletados
    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([data_hora, temperatura, umidade])

    # Salva o arquivo
    wb.save(caminho_arquivo)

    # Atualiza a interface gráfica com os dados coletados
    label_resultado.config(text=f"Temperatura: {temperatura}°C\nUmidade: {umidade}%")

# Cria a janela principal
janela_principal = tk.Tk()
janela_principal.title("Previsão do Tempo")

# Cria um botão para buscar a previsão
botao_buscar = tk.Button(janela_principal, text="Buscar previsão", command=buscar_previsao)
botao_buscar.pack(pady=20)

# Cria um label para exibir o resultado
label_resultado = tk.Label(janela_principal, text="")
label_resultado.pack(pady=20)

# Inicia a interface gráfica
janela_principal.mainloop()
